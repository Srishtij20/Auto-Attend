import io
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK = colors.HexColor("#1c1917")
ACCENT = colors.HexColor("#4f46e5")
GREEN = colors.HexColor("#16a34a")
RED = colors.HexColor("#dc2626")
AMBER = colors.HexColor("#d97706")
LIGHT = colors.HexColor("#fafaf9")
BORDER = colors.HexColor("#e7e5e4")

def generate_pdf(summary: List[Dict], date_str: str, title: str = "Attendance Report") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)

    title_style = ParagraphStyle("T", fontSize=20, fontName="Helvetica-Bold", textColor=DARK, spaceAfter=4)
    sub_style = ParagraphStyle("S", fontSize=10, fontName="Helvetica", textColor=colors.HexColor("#78716c"))

    present = [r for r in summary if r.get("status") == "present"]
    half = [r for r in summary if r.get("status") == "half_day"]
    absent = [r for r in summary if r.get("status") not in ("present", "half_day")]
    total = len(summary)
    pct = round(len(present) / total * 100, 1) if total else 0

    def fmt(iso):
        if not iso: return "—"
        try: return datetime.fromisoformat(str(iso)).strftime("%H:%M")
        except: return "—"

    elements = [
        Paragraph(title, title_style),
        Paragraph(f"Date: {date_str}  ·  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_style),
        Spacer(1, 10),
        HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=12),
    ]

    # Summary stats
    stats = [["Total", "Present", "Half Day", "Absent", "Attendance %"],
             [str(total), str(len(present)), str(len(half)), str(len(absent)), f"{pct}%"]]
    stats_tbl = Table(stats, colWidths=[3.4*cm]*5)
    stats_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), DARK), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ALIGN", (0,0), (-1,-1), "CENTER"), ("TOPPADDING", (0,0), (-1,-1), 8), ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("BACKGROUND", (1,1), (1,1), colors.HexColor("#dcfce7")), ("TEXTCOLOR", (1,1), (1,1), GREEN),
        ("BACKGROUND", (3,1), (3,1), colors.HexColor("#fee2e2")), ("TEXTCOLOR", (3,1), (3,1), RED),
        ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"), ("FONTSIZE", (0,1), (-1,1), 13),
        ("BOX", (0,0), (-1,-1), 0.5, BORDER), ("INNERGRID", (0,0), (-1,-1), 0.3, BORDER),
    ]))
    elements += [stats_tbl, Spacer(1, 18)]

    # Detail table
    rows = [["#", "Employee", "Department", "Check In", "Check Out", "Hours", "Status"]]
    for i, r in enumerate(sorted(summary, key=lambda x: x.get("employee_name", "")), 1):
        status = r.get("status", "absent")
        rows.append([
            str(i),
            r.get("employee_name", "—"),
            r.get("department") or "—",
            fmt(r.get("check_in")),
            fmt(r.get("check_out")),
            f"{r.get('total_hours', 0)}h" if r.get("total_hours") else "—",
            status.replace("_", " ").title(),
        ])

    col_w = [1*cm, 5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2.5*cm]
    detail_tbl = Table(rows, colWidths=col_w, repeatRows=1)
    style = [
        ("BACKGROUND", (0,0), (-1,0), DARK), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "CENTER"), ("ALIGN", (1,1), (2,-1), "LEFT"),
        ("TOPPADDING", (0,0), (-1,-1), 6), ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, LIGHT]),
        ("BOX", (0,0), (-1,-1), 0.5, BORDER), ("INNERGRID", (0,0), (-1,-1), 0.3, BORDER),
    ]
    for i, row in enumerate(rows[1:], 1):
        s = row[6].lower().replace(" ", "_")
        if s == "present":
            style += [("BACKGROUND",(6,i),(6,i),colors.HexColor("#dcfce7")),("TEXTCOLOR",(6,i),(6,i),GREEN)]
        elif s == "half_day":
            style += [("BACKGROUND",(6,i),(6,i),colors.HexColor("#fef3c7")),("TEXTCOLOR",(6,i),(6,i),AMBER)]
        else:
            style += [("BACKGROUND",(6,i),(6,i),colors.HexColor("#fee2e2")),("TEXTCOLOR",(6,i),(6,i),RED)]
    detail_tbl.setStyle(TableStyle(style))
    elements.append(detail_tbl)

    elements += [
        Spacer(1, 20), HRFlowable(width="100%", thickness=0.3, color=BORDER),
        Paragraph(f"Auto-Attend · {datetime.now().strftime('%d %b %Y %H:%M')} · Confidential",
                  ParagraphStyle("F", fontSize=8, textColor=colors.HexColor("#a8a29e"))),
    ]
    doc.build(elements)
    return buf.getvalue()


def generate_excel(summary: List[Dict], date_str: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    hdr_fill = PatternFill("solid", fgColor="1c1917")
    green_fill = PatternFill("solid", fgColor="dcfce7")
    red_fill = PatternFill("solid", fgColor="fee2e2")
    amber_fill = PatternFill("solid", fgColor="fef3c7")
    alt_fill = PatternFill("solid", fgColor="fafaf9")
    thin = Side(style="thin", color="e7e5e4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Attendance Report — {date_str}"
    ws["A1"].font = Font(bold=True, size=14, color="1c1917")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Employee", "Department", "Check In", "Check Out", "Hours", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 20

    def fmt(iso):
        if not iso: return "—"
        try: return datetime.fromisoformat(str(iso)).strftime("%H:%M")
        except: return "—"

    for i, r in enumerate(sorted(summary, key=lambda x: x.get("employee_name", "")), 1):
        row = 3 + i
        status = r.get("status", "absent")
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        sf = green_fill if status == "present" else amber_fill if status == "half_day" else red_fill

        vals = [i, r.get("employee_name","—"), r.get("department") or "—",
                fmt(r.get("check_in")), fmt(r.get("check_out")),
                f"{r.get('total_hours',0)}h" if r.get("total_hours") else "—",
                status.replace("_"," ").title()]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
            cell.border = border
            cell.fill = sf if col == 7 else fill
        ws.row_dimensions[row].height = 18

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()